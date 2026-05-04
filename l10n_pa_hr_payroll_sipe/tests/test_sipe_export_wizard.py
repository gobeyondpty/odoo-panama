# Part of Odoo. See LICENSE file for full copyright and licensing details.
"""Integration tests for the SIPE export wizard.

These tests use Odoo's TransactionCase. They build minimal hr.payslip
records, run the wizard, and verify the resulting XLSX bytes contain the
expected per-employee aggregation. No real payroll structures are loaded.
"""
import base64
from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from odoo.exceptions import UserError, ValidationError
from odoo.tests.common import TransactionCase, tagged

from odoo.addons.l10n_pa_hr_payroll_sipe.lib.sipe_writer import SIPE_COLUMNS


@tagged("post_install", "-at_install", "l10n_pa")
class TestSipeExportWizard(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.company = cls.env["res.company"].create({
            "name": "PA SIPE Test Co.",
            "country_id": cls.env.ref("base.pa").id,
        })
        cls.env.user.company_ids = [(4, cls.company.id)]
        cls.env.user.company_id = cls.company

        cls.employee = cls.env["hr.employee"].create({
            "name": "Juan Perez",
            "company_id": cls.company.id,
            "identification_id": "8-123-456",
            "l10n_pa_css_number": "1234567",
            "l10n_pa_sipe_id_type": "CEDULA",
        })
        cls.employee2 = cls.env["hr.employee"].create({
            "name": "Maria Lopez",
            "company_id": cls.company.id,
            "identification_id": "8-987-654",
            "l10n_pa_css_number": "7654321",
            "l10n_pa_sipe_id_type": "CEDULA",
        })

        # Tag the BASIC rule shipped by l10n_pa_hr_payroll so the wizard
        # has at least one rule to aggregate.
        cls.basic_rule = cls.env.ref(
            "l10n_pa_hr_payroll.l10n_pa_rule_basic", raise_if_not_found=False)
        if cls.basic_rule:
            # The data XML in this module already sets sueldo, but make it
            # explicit here so the test does not depend on data load order.
            cls.basic_rule.l10n_pa_sipe_column = "sueldo"

    def _make_payslip(self, employee, basic_amount, dfrom, dto, state="validated"):
        """Create a minimal posted payslip with a single BASIC line.

        We bypass the structure/contract machinery and write the line
        directly because the goal is to test the wizard's aggregation,
        not the payroll engine.
        """
        if not self.basic_rule:
            self.skipTest("l10n_pa_hr_payroll BASIC rule not loaded")
        payslip = self.env["hr.payslip"].create({
            "name": f"PS {employee.name} {dfrom}",
            "employee_id": employee.id,
            "company_id": self.company.id,
            "date_from": dfrom,
            "date_to": dto,
            "state": state,
        })
        self.env["hr.payslip.line"].create({
            "slip_id": payslip.id,
            "salary_rule_id": self.basic_rule.id,
            "code": "BASIC",
            "name": "Basic",
            "sequence": 1,
            "amount": basic_amount,
            "quantity": 1.0,
            "rate": 100.0,
            "total": basic_amount,
            "employee_id": employee.id,
        })
        return payslip

    def _read_xlsx(self, b64_data):
        wb = load_workbook(BytesIO(base64.b64decode(b64_data)), read_only=True)
        return [list(r) for r in wb.active.iter_rows(values_only=True)]

    def test_wizard_with_no_payslips_raises(self):
        wizard = self.env["l10n.pa.sipe.export.wizard"].create({
            "company_id": self.company.id,
            "date_start": date(2026, 5, 1),
            "date_end": date(2026, 5, 31),
        })
        with self.assertRaises(UserError):
            wizard.action_generate()

    def test_wizard_aggregates_two_employees(self):
        self._make_payslip(self.employee, 1000.0, date(2026, 5, 1), date(2026, 5, 31))
        self._make_payslip(self.employee2, 1500.0, date(2026, 5, 1), date(2026, 5, 31))

        wizard = self.env["l10n.pa.sipe.export.wizard"].create({
            "company_id": self.company.id,
            "date_start": date(2026, 5, 1),
            "date_end": date(2026, 5, 31),
        })
        wizard.action_generate()
        self.assertEqual(wizard.state, "done")
        self.assertTrue(wizard.file_data)
        self.assertTrue(wizard.file_name.endswith(".xlsx"))

        cells = self._read_xlsx(wizard.file_data)
        self.assertEqual(cells[0], list(SIPE_COLUMNS))
        # 2 employees, header + 2 data rows
        self.assertEqual(len(cells), 3)
        # Find rows by cédula (order is not guaranteed)
        by_cedula = {row[1]: row for row in cells[1:]}
        self.assertEqual(by_cedula["8-123-456"][5], 1000.0)
        self.assertEqual(by_cedula["8-987-654"][5], 1500.0)

    def test_wizard_sums_multiple_payslips_for_same_employee(self):
        # Two payslips in the same period — biweekly. They should sum.
        self._make_payslip(self.employee, 500.0, date(2026, 5, 1), date(2026, 5, 15))
        self._make_payslip(self.employee, 500.0, date(2026, 5, 16), date(2026, 5, 31))

        wizard = self.env["l10n.pa.sipe.export.wizard"].create({
            "company_id": self.company.id,
            "date_start": date(2026, 5, 1),
            "date_end": date(2026, 5, 31),
        })
        wizard.action_generate()
        cells = self._read_xlsx(wizard.file_data)
        self.assertEqual(len(cells), 2)
        self.assertEqual(cells[1][5], 1000.0)

    def test_wizard_excludes_draft_payslips(self):
        self._make_payslip(
            self.employee, 1000.0, date(2026, 5, 1), date(2026, 5, 31), state="draft",
        )

        wizard = self.env["l10n.pa.sipe.export.wizard"].create({
            "company_id": self.company.id,
            "date_start": date(2026, 5, 1),
            "date_end": date(2026, 5, 31),
        })
        with self.assertRaises(UserError):
            wizard.action_generate()

    def test_wizard_rejects_employee_without_css_number(self):
        bad_employee = self.env["hr.employee"].create({
            "name": "No CSS",
            "company_id": self.company.id,
            "identification_id": "8-555-555",
            # l10n_pa_css_number left empty
            "l10n_pa_sipe_id_type": "CEDULA",
        })
        self._make_payslip(bad_employee, 1000.0, date(2026, 5, 1), date(2026, 5, 31))

        wizard = self.env["l10n.pa.sipe.export.wizard"].create({
            "company_id": self.company.id,
            "date_start": date(2026, 5, 1),
            "date_end": date(2026, 5, 31),
        })
        # Don't use assertRaises: it triggers a transaction.flush() that
        # interacts badly with the partially-initialized employee record
        # in Odoo 19's hr_payroll. Catch manually.
        try:
            wizard.action_generate()
        except ValidationError as exc:
            self.assertIn("Seguro Social", str(exc))
        else:
            self.fail("Expected ValidationError for missing CSS number")

    def test_wizard_rejects_inverted_date_range(self):
        wizard = self.env["l10n.pa.sipe.export.wizard"].create({
            "company_id": self.company.id,
            "date_start": date(2026, 5, 31),
            "date_end": date(2026, 5, 1),
        })
        with self.assertRaises(ValidationError):
            wizard.action_generate()
