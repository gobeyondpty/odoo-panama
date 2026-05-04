# Part of Odoo. See LICENSE file for full copyright and licensing details.
"""Pure-Python tests for the SIPE XLSX writer.

These tests do not require an Odoo bootstrap. They verify column order,
header text, validation rules, and that openpyxl can read back what
write_sipe_xlsx produced. The expected column labels are the ones
documented in the official CSS Manual de Carga Masiva V2.
"""
from __future__ import annotations

import importlib.util
import math
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

# Load sipe_writer.py directly by file path so this test does not collide
# with the unrelated l10n_pa_hr_payroll/lib/ import that other repo-root
# tests already cache under the bare name `lib`.
_SIPE_WRITER_PATH = (
    Path(__file__).resolve().parents[1]
    / "l10n_pa_hr_payroll_sipe"
    / "lib"
    / "sipe_writer.py"
)
import sys  # noqa: E402

_spec = importlib.util.spec_from_file_location(
    "_sipe_writer_under_test", _SIPE_WRITER_PATH,
)
sipe_writer = importlib.util.module_from_spec(_spec)
# dataclasses(slots=True) inspects sys.modules[cls.__module__] during class
# creation, so register the module before exec.
sys.modules[_spec.name] = sipe_writer
_spec.loader.exec_module(sipe_writer)

FIELD_KEYS = sipe_writer.FIELD_KEYS
SIPE_COLUMNS = sipe_writer.SIPE_COLUMNS
SIPE_TIPO_CEDULA = sipe_writer.SIPE_TIPO_CEDULA
SIPE_TIPO_PASAPORTE = sipe_writer.SIPE_TIPO_PASAPORTE
SIPE_TIPO_SEGURO_SOCIAL = sipe_writer.SIPE_TIPO_SEGURO_SOCIAL
SIPE_VALID_TIPOS = sipe_writer.SIPE_VALID_TIPOS
SipeRow = sipe_writer.SipeRow
write_sipe_xlsx = sipe_writer.write_sipe_xlsx


class TestSipeWriterShape(unittest.TestCase):
    def test_25_columns(self):
        self.assertEqual(len(SIPE_COLUMNS), 25)
        self.assertEqual(len(FIELD_KEYS), 25)

    def test_first_column_is_tipo_documento(self):
        self.assertEqual(SIPE_COLUMNS[0], "Tipo de Documento")
        self.assertEqual(FIELD_KEYS[0], "tipo_documento")

    def test_last_column_is_indemnizacion(self):
        self.assertEqual(SIPE_COLUMNS[-1], "Indemnización")
        self.assertEqual(FIELD_KEYS[-1], "indemnizacion")

    def test_metadata_columns_first(self):
        self.assertEqual(
            FIELD_KEYS[:5],
            ("tipo_documento", "numero_documento", "numero_seguro_social",
             "nombre", "apellido"),
        )

    def test_unique_column_keys(self):
        self.assertEqual(len(set(FIELD_KEYS)), len(FIELD_KEYS))

    def test_unique_column_labels(self):
        self.assertEqual(len(set(SIPE_COLUMNS)), len(SIPE_COLUMNS))


class TestSipeRowValidation(unittest.TestCase):
    BASE = {
        "tipo_documento": SIPE_TIPO_CEDULA,
        "numero_documento": "8-123-456",
        "numero_seguro_social": "1234567",
        "nombre": "Juan",
        "apellido": "Perez",
    }

    def _row(self, **overrides):
        return SipeRow(**{**self.BASE, **overrides})

    def test_minimal_row_accepted(self):
        row = self._row()
        self.assertEqual(row.tipo_documento, SIPE_TIPO_CEDULA)
        self.assertEqual(row.sueldo, 0.0)

    def test_all_three_id_types_accepted(self):
        for tipo in SIPE_VALID_TIPOS:
            self._row(tipo_documento=tipo)

    def test_invalid_tipo_documento_rejected(self):
        with self.assertRaises(ValueError) as cm:
            self._row(tipo_documento="DNI")
        self.assertIn("tipo_documento", str(cm.exception))

    def test_empty_numero_documento_rejected(self):
        with self.assertRaises(ValueError):
            self._row(numero_documento="")

    def test_whitespace_only_nombre_rejected(self):
        with self.assertRaises(ValueError):
            self._row(nombre="   ")

    def test_negative_sueldo_rejected(self):
        with self.assertRaises(ValueError):
            self._row(sueldo=-100.0)

    def test_nan_sueldo_rejected(self):
        with self.assertRaises(ValueError):
            self._row(sueldo=float("nan"))

    def test_inf_sueldo_rejected(self):
        with self.assertRaises(ValueError):
            self._row(sueldo=math.inf)

    def test_bool_sueldo_rejected(self):
        # Python's bool subclasses int but we don't want True/False sneaking
        # in as a numeric value.
        with self.assertRaises(TypeError):
            self._row(sueldo=True)

    def test_string_sueldo_rejected(self):
        with self.assertRaises(TypeError):
            self._row(sueldo="100")

    def test_int_amount_normalized_to_float(self):
        row = self._row(sueldo=1000)
        self.assertIsInstance(row.sueldo, float)
        self.assertEqual(row.sueldo, 1000.0)


class TestSipeWriterRoundtrip(unittest.TestCase):
    def _round_trip(self, rows):
        data = write_sipe_xlsx(rows)
        wb = load_workbook(BytesIO(data), read_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]

    def test_header_row_matches_spec(self):
        rows = []
        cells = self._round_trip(rows)
        self.assertEqual(cells[0], list(SIPE_COLUMNS))

    def test_single_row_serialization(self):
        row = SipeRow(
            tipo_documento=SIPE_TIPO_CEDULA,
            numero_documento="8-123-456",
            numero_seguro_social="1234567",
            nombre="Juan",
            apellido="Perez",
            sueldo=1000.0,
            impuesto_sobre_renta=15.50,
            decimo_tercer_mes=83.33,
        )
        cells = self._round_trip([row])
        self.assertEqual(len(cells), 2)
        data_row = cells[1]
        self.assertEqual(data_row[0], "CEDULA")
        self.assertEqual(data_row[1], "8-123-456")
        self.assertEqual(data_row[2], "1234567")
        self.assertEqual(data_row[3], "Juan")
        self.assertEqual(data_row[4], "Perez")
        self.assertEqual(data_row[5], 1000.0)
        self.assertEqual(data_row[7], 15.5)
        self.assertEqual(data_row[8], 83.33)
        # Unset numeric columns should serialize as 0
        self.assertEqual(data_row[24], 0)

    def test_three_id_types_all_round_trip(self):
        rows = [
            SipeRow(SIPE_TIPO_CEDULA, "8-123-456", "1234567", "Juan", "Perez"),
            SipeRow(SIPE_TIPO_PASAPORTE, "PA12345", "2345678", "Anne", "Smith"),
            SipeRow(SIPE_TIPO_SEGURO_SOCIAL, "SS999", "3456789", "Pedro", "Lopez"),
        ]
        cells = self._round_trip(rows)
        self.assertEqual(len(cells), 4)
        self.assertEqual(cells[1][0], "CEDULA")
        self.assertEqual(cells[2][0], "PASAPORTE")
        self.assertEqual(cells[3][0], "SEGURO SOCIAL")

    def test_unicode_preserved(self):
        row = SipeRow(
            SIPE_TIPO_CEDULA, "8-123-456", "1234567",
            "José María", "Núñez Peña",
        )
        cells = self._round_trip([row])
        self.assertEqual(cells[1][3], "José María")
        self.assertEqual(cells[1][4], "Núñez Peña")

    def test_large_planilla_streams(self):
        # 2000 rows is a realistic upper bound for a single empleador group;
        # write_sipe_xlsx uses openpyxl write_only mode so this should not
        # OOM and should serialize in a reasonable time.
        rows = [
            SipeRow(
                SIPE_TIPO_CEDULA,
                f"8-{i:04d}-{i:04d}",
                f"{i:07d}",
                f"Empleado{i}",
                f"Apellido{i}",
                sueldo=1000.0,
            )
            for i in range(2000)
        ]
        data = write_sipe_xlsx(rows)
        self.assertGreater(len(data), 1000)
        wb = load_workbook(BytesIO(data), read_only=True)
        # Header + 2000 employee rows
        self.assertEqual(sum(1 for _ in wb.active.iter_rows()), 2001)


if __name__ == "__main__":
    unittest.main()
